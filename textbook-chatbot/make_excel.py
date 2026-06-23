#!/usr/bin/env python3
"""
Excel文書生成ツール
使い方:
  python3 ~/textbook-chatbot/make_excel.py input.json
  python3 ~/textbook-chatbot/make_excel.py input.json --out ~/Desktop/output.xlsx

JSON形式:
{
  "title": "タイトル",
  "sheets": [
    {
      "name": "シート名",
      "headers": ["列1", "列2", "列3"],
      "rows": [
        ["データA", "データB", "データC"],
        ["データD", "データE", "データF"]
      ],
      "note": "備考テキスト（任意）"
    }
  ],
  "references": ["参考文献1", "参考文献2"]
}
"""

import json, sys, argparse, re
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

OUTPUT_DIR = Path.home() / "Desktop"

# ── カラー定義 ──────────────────────────────────────────────────
COLOR_HEADER      = "1D4ED8"   # ヘッダー行（濃い青）
COLOR_HEADER_FONT = "FFFFFF"   # ヘッダー文字（白）
COLOR_ROW_ODD     = "F0F5FF"   # 奇数行（薄青）
COLOR_ROW_EVEN    = "FFFFFF"   # 偶数行（白）
COLOR_TITLE       = "1D4ED8"   # タイトル文字
COLOR_NOTE        = "666666"   # 備考文字
COLOR_REF         = "888888"   # 参考文献文字

THIN_BORDER = Border(
    left   = Side(style="thin", color="CCCCCC"),
    right  = Side(style="thin", color="CCCCCC"),
    top    = Side(style="thin", color="CCCCCC"),
    bottom = Side(style="thin", color="CCCCCC"),
)


def auto_col_width(ws, min_w: int = 8, max_w: int = 50):
    """列幅を内容に合わせて自動調整"""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                # 日本語は2文字分として計算
                val = str(cell.value)
                length = sum(2 if ord(c) > 127 else 1 for c in val)
                max_len = max(max_len, length)
        ws.column_dimensions[col_letter].width = max(min_w, min(max_len + 2, max_w))


def write_sheet(wb: openpyxl.Workbook, sheet_data: dict):
    """1シートを書き込む"""
    ws = wb.create_sheet(title=sheet_data.get("name", "シート"))
    headers  = sheet_data.get("headers", [])
    rows     = sheet_data.get("rows", [])
    note     = sheet_data.get("note", "")
    subtitle = sheet_data.get("subtitle", "")

    row_cursor = 1

    # サブタイトル（任意）
    if subtitle:
        cell = ws.cell(row=row_cursor, column=1, value=subtitle)
        cell.font = Font(bold=True, size=12, color=COLOR_TITLE)
        row_cursor += 1

    # 備考
    if note:
        cell = ws.cell(row=row_cursor, column=1, value=f"※ {note}")
        cell.font = Font(italic=True, size=9, color=COLOR_NOTE)
        row_cursor += 1

    if subtitle or note:
        row_cursor += 1   # 空行

    # ── ヘッダー行 ─────────────────────────────────────────────
    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row_cursor, column=c_idx, value=h)
        cell.font      = Font(bold=True, size=11, color=COLOR_HEADER_FONT)
        cell.fill      = PatternFill("solid", fgColor=COLOR_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER
    ws.row_dimensions[row_cursor].height = 22
    row_cursor += 1

    # ── データ行 ───────────────────────────────────────────────
    for r_idx, row_data in enumerate(rows):
        bg = COLOR_ROW_ODD if r_idx % 2 == 0 else COLOR_ROW_EVEN
        fill = PatternFill("solid", fgColor=bg)
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_cursor, column=c_idx, value=str(val) if val is not None else "")
            cell.font      = Font(size=10)
            cell.fill      = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border    = THIN_BORDER
        ws.row_dimensions[row_cursor].height = 18
        row_cursor += 1

    auto_col_width(ws)

    # 参考文献（シート末尾）
    refs = sheet_data.get("references", [])
    if refs:
        row_cursor += 1
        ws.cell(row=row_cursor, column=1, value="【参考文献】").font = Font(bold=True, size=9, color=COLOR_NOTE)
        row_cursor += 1
        for ref in refs:
            ws.cell(row=row_cursor, column=1, value=f"・{ref}").font = Font(size=9, color=COLOR_REF)
            row_cursor += 1

    # ウィンドウ枠の固定（ヘッダー行）
    ws.freeze_panes = ws.cell(row=(2 if not (subtitle or note) else 4), column=1)

    return ws


def generate_excel(data: dict, out_path: Path):
    """Excelファイルを生成"""
    wb = openpyxl.Workbook()
    # デフォルトシートは後で削除（シートが1枚以上になってから）
    default_ws = wb.active

    global_refs = data.get("references", [])
    sheets = data.get("sheets", [])

    # sheetsが空なら簡易モード（dataをそのままシートに）
    if not sheets and data.get("headers"):
        sheets = [{
            "name":    data.get("title", "データ"),
            "headers": data.get("headers", []),
            "rows":    data.get("rows", []),
            "note":    data.get("note", ""),
            "references": global_refs,
        }]

    for sh in sheets:
        if global_refs and not sh.get("references"):
            sh["references"] = global_refs
        write_sheet(wb, sh)

    # デフォルトシートを削除（他にシートがある場合）
    if len(wb.sheetnames) > 1 and default_ws.title in wb.sheetnames:
        wb.remove(default_ws)

    # 参考文献シート（全体）
    if global_refs and len(sheets) > 1:
        ws_ref = wb.create_sheet(title="参考文献")
        ws_ref.column_dimensions["A"].width = 80
        ws_ref.cell(row=1, column=1, value="参考文献").font = Font(bold=True, size=12, color=COLOR_TITLE)
        for i, ref in enumerate(global_refs, 2):
            ws_ref.cell(row=i, column=1, value=f"・{ref}").font = Font(size=10, color=COLOR_NOTE)

    wb.save(str(out_path))
    return out_path


def main():
    parser = argparse.ArgumentParser(description="Excel生成")
    parser.add_argument("input", help="入力JSONファイル（または - でstdin）")
    parser.add_argument("--out", "-o", default="", help="出力ファイルパス")
    args = parser.parse_args()

    if args.input == "-":
        data = json.load(sys.stdin)
    else:
        with open(args.input, encoding="utf-8") as f:
            data = json.load(f)

    if args.out:
        out_path = Path(args.out)
    else:
        ts   = datetime.now().strftime("%Y%m%d_%H%M")
        name = re.sub(r'[^\w぀-鿿]', '_', data.get("title", "output"))
        out_path = OUTPUT_DIR / f"{name}_{ts}.xlsx"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    result = generate_excel(data, out_path)
    print(f"✅ Excel保存完了: {result}")


if __name__ == "__main__":
    main()
